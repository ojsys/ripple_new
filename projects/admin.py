from django.contrib import admin
from django.utils.safestring import mark_safe
from django.db import models
from django.urls import reverse
from django.http import HttpResponseRedirect, HttpResponse
from django_json_widget.widgets import JSONEditorWidget
from django.contrib.auth.admin import UserAdmin
from django.utils.html import format_html
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
import io

# Import only the models that definitely exist
from apps.accounts.models import CustomUser, FounderProfile, InvestorProfile, PartnerProfile, RegistrationPayment, PendingRegistration
from .models import (
    Category, Project, FundingType,
    Reward, Pledge, InvestmentTerm, Investment
)

# Try to import optional models
try:
    from .models import Update
    HAS_UPDATE = True
except ImportError:
    HAS_UPDATE = False

class ExcelExportMixin:
    """Mixin to add Excel export functionality to admin classes"""
    
    def export_to_excel(self, request, queryset):
        """Export selected objects to Excel"""
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{self.model._meta.verbose_name_plural}"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Get field names and verbose names
        field_names = self.get_export_fields()
        verbose_names = [self.get_field_verbose_name(field) for field in field_names]
        
        # Write headers
        for col, header in enumerate(verbose_names, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write data
        for row, obj in enumerate(queryset, 2):
            for col, field_name in enumerate(field_names, 1):
                value = self.get_field_value(obj, field_name)
                ws.cell(row=row, column=col, value=value)
        
        # Auto-adjust column widths
        for col in range(1, len(field_names) + 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].auto_size = True
            # Set minimum width
            ws.column_dimensions[column_letter].width = max(ws.column_dimensions[column_letter].width, 12)
        
        # Create HTTP response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{self.model._meta.verbose_name_plural}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        # Save workbook to response
        wb.save(response)
        return response
    
    def get_export_fields(self):
        """Get fields to export - override in subclasses for custom fields"""
        return [field.name for field in self.model._meta.fields if not field.name.endswith('_ptr')]
    
    def get_field_verbose_name(self, field_name):
        """Get verbose name for field"""
        try:
            field = self.model._meta.get_field(field_name)
            return field.verbose_name.title()
        except:
            return field_name.replace('_', ' ').title()
    
    def get_field_value(self, obj, field_name):
        """Get field value with proper formatting"""
        try:
            # Handle related field lookups
            if '__' in field_name:
                parts = field_name.split('__')
                value = obj
                for part in parts:
                    value = getattr(value, part, None)
                    if value is None:
                        return ""
                return str(value)
            
            value = getattr(obj, field_name)
            
            # Handle different field types
            if value is None:
                return ""
            elif hasattr(value, 'all'):  # ManyToMany field
                return ", ".join(str(item) for item in value.all())
            elif hasattr(value, '__call__'):  # Method
                return str(value())
            elif isinstance(value, bool):
                return "Yes" if value else "No"
            elif isinstance(value, (datetime, type(datetime.now().date()))):
                return value.strftime("%Y-%m-%d %H:%M:%S") if hasattr(value, 'hour') else value.strftime("%Y-%m-%d")
            else:
                return str(value)
        except:
            return ""
    
    export_to_excel.short_description = "Export selected items to Excel"

# Simple admin registrations with Excel export
@admin.register(Category)
class CategoryAdmin(admin.ModelAdmin, ExcelExportMixin):
    list_display = ('name',)
    search_fields = ('name',)
    actions = ['export_to_excel']
    
    def get_export_fields(self):
        return ['name']

@admin.register(FundingType)
class FundingTypeAdmin(admin.ModelAdmin, ExcelExportMixin):
    list_display = ('name', 'description')
    search_fields = ('name',)
    actions = ['export_to_excel']
    
    def get_export_fields(self):
        return ['name', 'description']

@admin.register(Project)
class ProjectAdmin(admin.ModelAdmin, ExcelExportMixin):
    list_display = ('title', 'creator', 'category', 'funding_goal', 'amount_raised', 'status')
    list_filter = ('category', 'status', 'funding_type', 'created_at')
    search_fields = ('title', 'creator__email')
    actions = ['export_to_excel']
    
    def get_export_fields(self):
        return ['title', 'creator__email', 'creator__username', 'category__name', 'funding_type__name', 
                'funding_goal', 'amount_raised', 'total_investment_raised', 'description', 'short_description', 
                'location', 'video_url', 'status', 'admin_notes', 'created_at', 'deadline']

@admin.register(Reward)
class RewardAdmin(admin.ModelAdmin, ExcelExportMixin):
    list_display = ('title', 'project', 'amount')
    list_filter = ('project',)
    search_fields = ('title', 'project__title')
    actions = ['export_to_excel']
    
    def get_export_fields(self):
        return ['title', 'project__title', 'description', 'amount']

@admin.register(Pledge)
class PledgeAdmin(admin.ModelAdmin, ExcelExportMixin):
    list_display = ('backer', 'project', 'amount', 'pledged_at')
    list_filter = ('pledged_at', 'project')
    search_fields = ('backer__email', 'project__title')
    actions = ['export_to_excel']
    
    def get_export_fields(self):
        return ['backer__email', 'backer__username', 'project__title', 'amount', 'reward__title', 'pledged_at']

@admin.register(InvestmentTerm)
class InvestmentTermAdmin(admin.ModelAdmin, ExcelExportMixin):
    list_display = ('project', 'equity_offered', 'minimum_investment', 'valuation')
    list_filter = ('project',)
    search_fields = ('project__title',)
    actions = ['export_to_excel']
    
    def get_export_fields(self):
        return ['project__title', 'equity_offered', 'minimum_investment', 'maximum_investment', 'valuation', 'deadline']

@admin.register(Investment)
class InvestmentAdmin(admin.ModelAdmin, ExcelExportMixin):
    list_display = ('investor', 'project', 'amount', 'status', 'created_at')
    list_filter = ('status', 'created_at', 'project')
    search_fields = ('investor__email', 'project__title')
    actions = ['export_to_excel']
    
    def get_export_fields(self):
        return ['investor__email', 'investor__username', 'project__title', 'amount', 'equity_percentage', 
                'actual_return', 'terms__equity_offered', 'status', 'is_counted', 'created_at']

# Register optional models if they exist
if HAS_UPDATE:
    @admin.register(Update)
    class UpdateAdmin(admin.ModelAdmin, ExcelExportMixin):
        list_display = ('title', 'project', 'created_at')
        list_filter = ('project', 'created_at')
        search_fields = ('title', 'project__title')
        actions = ['export_to_excel']
        
        def get_export_fields(self):
            return ['title', 'project__title', 'content', 'created_at']




