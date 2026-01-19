from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
from django.core.paginator import Paginator
from django.db.models import Sum, Q
from django.utils import timezone
from django.conf import settings
from decimal import Decimal
import uuid
import requests
import csv
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .email_utils import (
    send_token_purchase_confirmation_to_user,
    send_token_purchase_notification_to_admin,
    send_withdrawal_request_to_user,
    send_withdrawal_request_to_admin,
)

from .models import (
    TokenPackage, PartnerCapitalAccount, Venture,
    VentureInvestment, SRTTransaction, TokenPurchase, TokenWithdrawal
)
from .forms import WithdrawalForm, InvestmentModifyForm


def partner_required(view_func):
    """Decorator to ensure user is an SRT partner (either by user_type or is_srt_partner flag)"""
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            messages.error(request, "Please log in to access the partner dashboard.")
            return redirect('accounts:login')
        # Check if user is an SRT partner (either by user_type or is_srt_partner flag)
        is_partner = request.user.user_type == 'partner' or getattr(request.user, 'is_srt_partner', False)
        if not is_partner:
            messages.info(request, "Become an SRT Partner to access investment opportunities.")
            return redirect('srt:become_partner')
        return view_func(request, *args, **kwargs)
    return wrapper


def get_or_create_capital_account(user):
    """Get or create a partner's capital account"""
    account, created = PartnerCapitalAccount.objects.get_or_create(partner=user)
    return account


@login_required
def become_partner(request):
    """View for users to become an SRT Partner"""
    # Check if already a partner
    is_partner = request.user.user_type == 'partner' or getattr(request.user, 'is_srt_partner', False)
    if is_partner:
        return redirect('srt:dashboard')

    if request.method == 'POST':
        # Process the partner registration
        agree_terms = request.POST.get('agree_terms')

        if not agree_terms:
            messages.error(request, "You must agree to the terms and conditions.")
            return redirect('srt:become_partner')

        # Set user as SRT partner
        request.user.is_srt_partner = True
        request.user.srt_partner_since = timezone.now()
        request.user.save()

        # Create PartnerProfile if it doesn't exist
        from apps.accounts.models import PartnerProfile
        partner_profile, created = PartnerProfile.objects.get_or_create(
            user=request.user,
            defaults={
                'partner_id': f"SRT-{uuid.uuid4().hex[:8].upper()}",
                'accreditation_status': 'pending',
            }
        )

        # Create PartnerCapitalAccount
        get_or_create_capital_account(request.user)

        messages.success(
            request,
            "Welcome to SRT! You are now an SRT Partner. You can purchase tokens and invest in ventures."
        )
        return redirect('srt:dashboard')

    context = {
        'user': request.user,
    }
    return render(request, 'srt/become_partner.html', context)


@login_required
@partner_required
def partner_dashboard(request):
    """Main partner dashboard view"""
    account = get_or_create_capital_account(request.user)

    # Get partner profile
    try:
        partner_profile = request.user.partner_profile
    except:
        partner_profile = None

    # Recent transactions
    recent_transactions = SRTTransaction.objects.filter(
        account=account
    ).select_related('venture')[:5]

    # Active investments
    active_investments = VentureInvestment.objects.filter(
        partner=request.user,
        status__in=['active', 'pending']
    ).select_related('venture')[:5]

    # Featured ventures
    featured_ventures = Venture.objects.filter(
        status='open',
        is_featured=True
    )[:3]

    # Stats
    total_invested = VentureInvestment.objects.filter(
        partner=request.user,
        status='active'
    ).aggregate(total=Sum('tokens_invested'))['total'] or 0

    active_investment_count = VentureInvestment.objects.filter(
        partner=request.user,
        status='active'
    ).count()

    # Expected returns
    expected_returns = VentureInvestment.objects.filter(
        partner=request.user,
        status='active'
    ).aggregate(total=Sum('expected_return'))['total'] or 0

    # Total returns earned (from matured investments)
    total_returns_earned = VentureInvestment.objects.filter(
        partner=request.user,
        status='matured'
    ).aggregate(total=Sum('actual_return'))['total'] or 0

    # Upcoming maturities (next 30 days)
    from datetime import timedelta
    thirty_days = timezone.now().date() + timedelta(days=30)
    upcoming_maturities = VentureInvestment.objects.filter(
        partner=request.user,
        status='active',
        maturity_date__lte=thirty_days
    ).select_related('venture').order_by('maturity_date')[:3]

    context = {
        'account': account,
        'partner_profile': partner_profile,
        'recent_transactions': recent_transactions,
        'active_investments': active_investments,
        'featured_ventures': featured_ventures,
        'total_invested': total_invested,
        'active_investment_count': active_investment_count,
        'expected_returns': expected_returns,
        'total_returns_earned': total_returns_earned,
        'upcoming_maturities': upcoming_maturities,
    }
    return render(request, 'srt/dashboard.html', context)


@login_required
@partner_required
def buy_tokens(request):
    """Token purchase page"""
    account = get_or_create_capital_account(request.user)

    # Token limits (monthly)
    MAX_TOKENS_PER_MONTH = 200
    TOKEN_RATE = Decimal('2000')  # 1 SRT = ₦2,000

    # Calculate tokens purchased this month
    from datetime import datetime
    current_month_start = timezone.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    tokens_purchased_this_month = TokenPurchase.objects.filter(
        partner=request.user,
        status='completed',
        created_at__gte=current_month_start
    ).aggregate(total=Sum('tokens'))['total'] or 0

    # Calculate how many tokens the partner can still buy this month
    tokens_remaining = max(0, MAX_TOKENS_PER_MONTH - int(tokens_purchased_this_month))

    context = {
        'account': account,
        'max_tokens': MAX_TOKENS_PER_MONTH,
        'tokens_purchased_this_month': int(tokens_purchased_this_month),
        'tokens_remaining': tokens_remaining,
        'token_rate': TOKEN_RATE,
        'current_month': timezone.now().strftime('%B %Y'),
        'paystack_public_key': settings.PAYSTACK_PUBLIC_KEY,
    }
    return render(request, 'srt/buy_tokens.html', context)


@login_required
@partner_required
@require_POST
def initialize_token_purchase(request):
    """Initialize Paystack payment for token purchase"""
    tokens_input = request.POST.get('tokens')

    # Token limits and rate (monthly)
    MAX_TOKENS_PER_MONTH = 200
    TOKEN_RATE = Decimal('2000')  # 1 SRT = ₦2,000

    account = get_or_create_capital_account(request.user)

    # Calculate tokens purchased this month
    current_month_start = timezone.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    tokens_purchased_this_month = TokenPurchase.objects.filter(
        partner=request.user,
        status='completed',
        created_at__gte=current_month_start
    ).aggregate(total=Sum('tokens'))['total'] or 0

    # Calculate tokens remaining this month
    tokens_remaining = MAX_TOKENS_PER_MONTH - int(tokens_purchased_this_month)

    if tokens_remaining <= 0:
        return JsonResponse({'error': 'You have reached your monthly token limit of 200 SRT. Please try again next month.'}, status=400)

    try:
        tokens = int(tokens_input)
        if tokens < 1:
            return JsonResponse({'error': 'Please enter at least 1 token.'}, status=400)
        if tokens > tokens_remaining:
            return JsonResponse({'error': f'You can only purchase up to {tokens_remaining} more tokens.'}, status=400)

        # Calculate amount (1 SRT = ₦2,000)
        amount_ngn = Decimal(tokens) * TOKEN_RATE
        amount_usd = amount_ngn / Decimal('1600')  # Exchange rate approximation
    except (ValueError, TypeError):
        return JsonResponse({'error': 'Invalid token amount.'}, status=400)

    # Generate reference
    reference = f"SRT-{uuid.uuid4().hex[:12].upper()}"

    # Create pending purchase (no bonus tokens)
    purchase = TokenPurchase.objects.create(
        partner=request.user,
        account=account,
        package=None,
        tokens=Decimal(tokens),
        bonus_tokens=Decimal('0'),
        amount_ngn=amount_ngn,
        amount_usd=amount_usd,
        paystack_reference=reference
    )

    # Initialize Paystack transaction
    headers = {
        'Authorization': f'Bearer {settings.PAYSTACK_SECRET_KEY}',
        'Content-Type': 'application/json',
    }

    data = {
        'email': request.user.email,
        'amount': int(amount_ngn * 100),  # Paystack uses kobo
        'reference': reference,
        'callback_url': request.build_absolute_uri('/srt/token-purchase-callback/'),
        'metadata': {
            'purchase_id': purchase.id,
            'tokens': str(tokens),
        }
    }

    try:
        response = requests.post(
            'https://api.paystack.co/transaction/initialize',
            json=data,
            headers=headers
        )
        result = response.json()

        if result.get('status'):
            return JsonResponse({
                'authorization_url': result['data']['authorization_url'],
                'reference': reference
            })
        else:
            purchase.status = 'failed'
            purchase.save()
            return JsonResponse({'error': result.get('message', 'Payment initialization failed')}, status=400)
    except Exception as e:
        purchase.status = 'failed'
        purchase.save()
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def token_purchase_callback(request):
    """Handle Paystack callback for token purchases"""
    reference = request.GET.get('reference')

    if not reference:
        messages.error(request, "Invalid payment reference.")
        return redirect('srt:buy_tokens')

    try:
        purchase = TokenPurchase.objects.get(paystack_reference=reference)
    except TokenPurchase.DoesNotExist:
        messages.error(request, "Purchase not found.")
        return redirect('srt:buy_tokens')

    # Verify with Paystack
    headers = {
        'Authorization': f'Bearer {settings.PAYSTACK_SECRET_KEY}',
    }

    try:
        response = requests.get(
            f'https://api.paystack.co/transaction/verify/{reference}',
            headers=headers
        )
        result = response.json()

        if result.get('status') and result['data']['status'] == 'success':
            # Complete the purchase
            if purchase.complete_purchase():
                messages.success(
                    request,
                    f"Payment successful! {purchase.total_tokens} SRT tokens have been added to your account."
                )
                # Send email notifications
                send_token_purchase_confirmation_to_user(purchase)
                send_token_purchase_notification_to_admin(purchase)
            else:
                messages.info(request, "This purchase has already been processed.")
        else:
            purchase.status = 'failed'
            purchase.save()
            messages.error(request, "Payment verification failed. Please contact support.")
    except Exception as e:
        messages.error(request, f"Error verifying payment: {str(e)}")

    return redirect('srt:dashboard')


@login_required
@partner_required
def venture_list(request):
    """List all available ventures"""
    ventures = Venture.objects.filter(status='open')

    # Filters
    risk_level = request.GET.get('risk')
    stage = request.GET.get('stage')
    industry = request.GET.get('industry')

    if risk_level:
        ventures = ventures.filter(risk_level=risk_level)
    if stage:
        ventures = ventures.filter(stage=stage)
    if industry:
        ventures = ventures.filter(industry__icontains=industry)

    # Get unique industries for filter
    industries = Venture.objects.filter(
        status='open'
    ).values_list('industry', flat=True).distinct()

    # Pagination
    paginator = Paginator(ventures, 12)
    page = request.GET.get('page', 1)
    ventures = paginator.get_page(page)

    account = get_or_create_capital_account(request.user)

    context = {
        'ventures': ventures,
        'industries': industries,
        'account': account,
        'current_risk': risk_level,
        'current_stage': stage,
        'current_industry': industry,
    }
    return render(request, 'srt/venture_list.html', context)


@login_required
@partner_required
def venture_detail(request, slug):
    """Venture detail page"""
    venture = get_object_or_404(Venture, slug=slug)
    account = get_or_create_capital_account(request.user)

    # Check if user already invested
    existing_investment = VentureInvestment.objects.filter(
        partner=request.user,
        venture=venture,
        status__in=['pending', 'active']
    ).first()

    # Recent investors (anonymized)
    recent_investments = VentureInvestment.objects.filter(
        venture=venture,
        status='active'
    ).order_by('-created_at')[:5]

    context = {
        'venture': venture,
        'account': account,
        'existing_investment': existing_investment,
        'recent_investments': recent_investments,
    }
    return render(request, 'srt/venture_detail.html', context)


@login_required
@partner_required
@require_POST
def invest_in_venture(request, slug):
    """Process investment in a venture"""
    venture = get_object_or_404(Venture, slug=slug, status='open')
    account = get_or_create_capital_account(request.user)

    try:
        amount = Decimal(request.POST.get('amount', 0))
    except:
        messages.error(request, "Invalid investment amount.")
        return redirect('srt:venture_detail', slug=slug)

    # Validate investment
    can_invest, message = venture.can_invest(amount)
    if not can_invest:
        messages.error(request, message)
        return redirect('srt:venture_detail', slug=slug)

    # Check available tokens
    if amount > account.available_tokens:
        messages.error(
            request,
            f"Insufficient tokens. You have {account.available_tokens:.2f} SRT available."
        )
        return redirect('srt:venture_detail', slug=slug)

    # Check if already invested
    existing = VentureInvestment.objects.filter(
        partner=request.user,
        venture=venture,
        status__in=['pending', 'active']
    ).exists()

    if existing:
        messages.warning(request, "You already have an active investment in this venture.")
        return redirect('srt:venture_detail', slug=slug)

    # Create investment
    investment = VentureInvestment.objects.create(
        partner=request.user,
        venture=venture,
        account=account,
        tokens_invested=amount,
        status='active'
    )

    # Lock tokens in account
    account.invest_tokens(amount, venture, f"Investment in {venture.title}")

    # Update venture amount raised
    venture.amount_raised += amount
    if venture.is_fully_funded:
        venture.status = 'funded'
    venture.save()

    messages.success(
        request,
        f"Investment successful! You've invested {amount:.2f} SRT in {venture.title}."
    )
    return redirect('srt:my_investments')


@login_required
@partner_required
def my_investments(request):
    """View all partner investments"""
    account = get_or_create_capital_account(request.user)

    # Filter by status
    status_filter = request.GET.get('status', '')

    investments = VentureInvestment.objects.filter(
        partner=request.user
    ).select_related('venture')

    if status_filter:
        investments = investments.filter(status=status_filter)

    # Stats
    stats = {
        'total_invested': investments.filter(
            status='active'
        ).aggregate(total=Sum('tokens_invested'))['total'] or 0,
        'expected_returns': investments.filter(
            status='active'
        ).aggregate(total=Sum('expected_return'))['total'] or 0,
        'active_count': investments.filter(status='active').count(),
        'matured_count': investments.filter(status='matured').count(),
    }

    # Pagination
    paginator = Paginator(investments, 10)
    page = request.GET.get('page', 1)
    investments = paginator.get_page(page)

    context = {
        'investments': investments,
        'account': account,
        'stats': stats,
        'current_status': status_filter,
    }
    return render(request, 'srt/my_investments.html', context)


@login_required
@partner_required
def investment_detail(request, reference):
    """View single investment details"""
    investment = get_object_or_404(
        VentureInvestment,
        reference=reference,
        partner=request.user
    )
    account = get_or_create_capital_account(request.user)

    # Related transactions
    transactions = SRTTransaction.objects.filter(
        account=account,
        venture=investment.venture
    ).order_by('-created_at')

    context = {
        'investment': investment,
        'account': account,
        'transactions': transactions,
    }
    return render(request, 'srt/investment_detail.html', context)


@login_required
@partner_required
def transaction_history(request):
    """View all token transactions"""
    account = get_or_create_capital_account(request.user)

    # Filter by type
    tx_type = request.GET.get('type', '')

    transactions = SRTTransaction.objects.filter(
        account=account
    ).select_related('venture', 'token_package')

    if tx_type:
        transactions = transactions.filter(transaction_type=tx_type)

    # Pagination
    paginator = Paginator(transactions, 20)
    page = request.GET.get('page', 1)
    transactions = paginator.get_page(page)

    context = {
        'transactions': transactions,
        'account': account,
        'current_type': tx_type,
        'transaction_types': SRTTransaction.TRANSACTION_TYPES,
    }
    return render(request, 'srt/transaction_history.html', context)


@login_required
@partner_required
def partner_profile(request):
    """Partner profile settings"""
    account = get_or_create_capital_account(request.user)

    try:
        partner_profile = request.user.partner_profile
    except:
        partner_profile = None

    context = {
        'account': account,
        'partner_profile': partner_profile,
    }
    return render(request, 'srt/partner_profile.html', context)


@login_required
@partner_required
def withdraw_tokens(request):
    """Request token withdrawal"""
    account = get_or_create_capital_account(request.user)

    # Get pending withdrawals
    pending_withdrawals = TokenWithdrawal.objects.filter(
        partner=request.user,
        status__in=['pending', 'approved', 'processing']
    )

    # Recent completed withdrawals
    recent_withdrawals = TokenWithdrawal.objects.filter(
        partner=request.user,
        status__in=['completed', 'rejected', 'cancelled']
    )[:5]

    if request.method == 'POST':
        form = WithdrawalForm(request.POST, account=account)
        if form.is_valid():
            # Check if there's already a pending withdrawal
            if pending_withdrawals.exists():
                messages.warning(
                    request,
                    'You already have a pending withdrawal request. Please wait for it to be processed.'
                )
                return redirect('srt:withdraw_tokens')

            # Create withdrawal request
            withdrawal = form.save(commit=False)
            withdrawal.partner = request.user
            withdrawal.account = account
            withdrawal.save()  # This will calculate fee and amount_ngn

            # Send email notifications
            send_withdrawal_request_to_user(withdrawal)
            send_withdrawal_request_to_admin(withdrawal)

            messages.success(
                request,
                f'Withdrawal request for {withdrawal.tokens:.2f} SRT submitted! A 4% fee of ₦{withdrawal.fee:,.2f} has been applied. You will receive ₦{withdrawal.amount_ngn:,.2f}.'
            )
            return redirect('srt:withdrawal_detail', reference=withdrawal.reference)
    else:
        form = WithdrawalForm(account=account)

    context = {
        'form': form,
        'account': account,
        'pending_withdrawals': pending_withdrawals,
        'recent_withdrawals': recent_withdrawals,
    }
    return render(request, 'srt/withdraw_tokens.html', context)


@login_required
@partner_required
def withdrawal_detail(request, reference):
    """View withdrawal request details"""
    withdrawal = get_object_or_404(
        TokenWithdrawal,
        reference=reference,
        partner=request.user
    )
    account = get_or_create_capital_account(request.user)

    context = {
        'withdrawal': withdrawal,
        'account': account,
    }
    return render(request, 'srt/withdrawal_detail.html', context)


@login_required
@partner_required
@require_POST
def cancel_withdrawal(request, reference):
    """Cancel a pending withdrawal request"""
    withdrawal = get_object_or_404(
        TokenWithdrawal,
        reference=reference,
        partner=request.user,
        status='pending'
    )

    if withdrawal.cancel():
        messages.success(request, 'Withdrawal request cancelled successfully.')
    else:
        messages.error(request, 'Unable to cancel this withdrawal request.')

    return redirect('srt:withdraw_tokens')


@login_required
@partner_required
def my_withdrawals(request):
    """View all withdrawal requests"""
    account = get_or_create_capital_account(request.user)

    # Filter by status
    status_filter = request.GET.get('status', '')

    withdrawals = TokenWithdrawal.objects.filter(partner=request.user)

    if status_filter:
        withdrawals = withdrawals.filter(status=status_filter)

    # Stats
    stats = {
        'total_withdrawn': TokenWithdrawal.objects.filter(
            partner=request.user,
            status='completed'
        ).aggregate(total=Sum('tokens'))['total'] or 0,
        'pending_count': TokenWithdrawal.objects.filter(
            partner=request.user,
            status='pending'
        ).count(),
        'completed_count': TokenWithdrawal.objects.filter(
            partner=request.user,
            status='completed'
        ).count(),
    }

    # Pagination
    paginator = Paginator(withdrawals, 10)
    page = request.GET.get('page', 1)
    withdrawals = paginator.get_page(page)

    context = {
        'withdrawals': withdrawals,
        'account': account,
        'stats': stats,
        'current_status': status_filter,
    }
    return render(request, 'srt/my_withdrawals.html', context)


@login_required
@partner_required
def modify_investment(request, reference):
    """Modify an existing investment (early withdrawal)"""
    investment = get_object_or_404(
        VentureInvestment,
        reference=reference,
        partner=request.user,
        status='active'
    )
    account = get_or_create_capital_account(request.user)

    if request.method == 'POST':
        form = InvestmentModifyForm(request.POST, investment=investment)
        if form.is_valid():
            action = form.cleaned_data['action']
            reason = form.cleaned_data['reason']
            notes = form.cleaned_data.get('additional_notes', '')

            if action == 'full_withdraw':
                # Calculate penalty (10% of invested amount)
                penalty = investment.tokens_invested * Decimal('0.10')
                return_amount = investment.tokens_invested - penalty

                # Update investment status
                investment.status = 'withdrawn'
                investment.actual_return = return_amount
                investment.notes = f"Early withdrawal. Reason: {reason}. {notes}"
                investment.save()

                # Release tokens with penalty
                account.locked_tokens -= investment.tokens_invested
                account.token_balance += return_amount
                account.save()

                # Record transaction
                SRTTransaction.objects.create(
                    account=account,
                    transaction_type='return',
                    amount=return_amount,
                    balance_after=account.token_balance,
                    venture=investment.venture,
                    reference=f"EW-{uuid.uuid4().hex[:10].upper()}",
                    description=f"Early withdrawal from {investment.venture.title} (10% penalty applied)"
                )

                # Update venture
                investment.venture.amount_raised -= investment.tokens_invested
                investment.venture.save()

                messages.success(
                    request,
                    f'Investment withdrawn. {return_amount:.2f} SRT returned (10% penalty: {penalty:.2f} SRT).'
                )
            else:
                # Partial withdrawal
                amount = form.cleaned_data['amount']
                penalty = amount * Decimal('0.10')
                return_amount = amount - penalty

                # Update investment
                investment.tokens_invested -= amount
                investment.notes = f"Partial withdrawal of {amount} SRT. Reason: {reason}. {notes}"
                investment.save()

                # Update account
                account.locked_tokens -= amount
                account.token_balance += return_amount
                account.save()

                # Record transaction
                SRTTransaction.objects.create(
                    account=account,
                    transaction_type='return',
                    amount=return_amount,
                    balance_after=account.token_balance,
                    venture=investment.venture,
                    reference=f"PW-{uuid.uuid4().hex[:10].upper()}",
                    description=f"Partial withdrawal from {investment.venture.title} (10% penalty applied)"
                )

                # Update venture
                investment.venture.amount_raised -= amount
                investment.venture.save()

                messages.success(
                    request,
                    f'Partial withdrawal successful. {return_amount:.2f} SRT returned (10% penalty: {penalty:.2f} SRT).'
                )

            return redirect('srt:my_investments')
    else:
        form = InvestmentModifyForm(investment=investment)

    # Calculate potential returns
    penalty_rate = Decimal('0.10')
    potential_return = investment.tokens_invested * (1 - penalty_rate)

    context = {
        'form': form,
        'investment': investment,
        'account': account,
        'penalty_rate': penalty_rate * 100,
        'potential_return': potential_return,
    }
    return render(request, 'srt/modify_investment.html', context)


# ============================================
# EXPORT FUNCTIONS
# ============================================

def create_excel_workbook(title, headers, data, column_widths=None):
    """Helper function to create styled Excel workbook"""
    wb = Workbook()
    ws = wb.active
    ws.title = title

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Write data
    for row_num, row_data in enumerate(data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    # Set column widths
    if column_widths:
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
    else:
        # Auto-fit columns (approximate)
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

    return wb


@login_required
@partner_required
def export_my_withdrawals_excel(request):
    """Export user's withdrawals to Excel"""
    withdrawals = TokenWithdrawal.objects.filter(partner=request.user).order_by('-created_at')

    headers = [
        'Reference', 'Tokens (SRT)', 'Fee (NGN)', 'Amount (NGN)',
        'Bank Name', 'Account Number', 'Account Name',
        'Status', 'Created Date', 'Processed Date', 'Completed Date'
    ]

    data = []
    for w in withdrawals:
        data.append([
            w.reference,
            float(w.tokens),
            float(w.fee) if w.fee else 0,
            float(w.amount_ngn),
            w.get_bank_name_display(),
            w.account_number,
            w.account_name,
            w.get_status_display(),
            w.created_at.strftime('%Y-%m-%d %H:%M') if w.created_at else '',
            w.processed_at.strftime('%Y-%m-%d %H:%M') if w.processed_at else '',
            w.completed_at.strftime('%Y-%m-%d %H:%M') if w.completed_at else '',
        ])

    wb = create_excel_workbook(
        'My Withdrawals',
        headers,
        data,
        [15, 15, 15, 15, 25, 15, 25, 15, 18, 18, 18]
    )

    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"my_withdrawals_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Save workbook to response
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response.write(buffer.getvalue())

    return response


@login_required
@partner_required
def export_my_transactions_excel(request):
    """Export user's transactions to Excel"""
    account = get_or_create_capital_account(request.user)
    transactions = SRTTransaction.objects.filter(account=account).order_by('-created_at')

    headers = [
        'Reference', 'Type', 'Amount (SRT)', 'Balance After',
        'Venture', 'Description', 'Payment Reference', 'Date'
    ]

    data = []
    for t in transactions:
        data.append([
            t.reference,
            t.get_transaction_type_display(),
            float(t.amount),
            float(t.balance_after),
            t.venture.title if t.venture else '',
            t.description or '',
            t.payment_reference or '',
            t.created_at.strftime('%Y-%m-%d %H:%M') if t.created_at else '',
        ])

    wb = create_excel_workbook(
        'My Transactions',
        headers,
        data,
        [18, 18, 15, 15, 25, 40, 20, 18]
    )

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"my_transactions_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response.write(buffer.getvalue())

    return response


@login_required
@partner_required
def export_my_transactions_csv(request):
    """Export user's transactions to CSV"""
    account = get_or_create_capital_account(request.user)
    transactions = SRTTransaction.objects.filter(account=account).order_by('-created_at')

    response = HttpResponse(content_type='text/csv')
    filename = f"my_transactions_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)
    writer.writerow([
        'Reference', 'Type', 'Amount (SRT)', 'Balance After',
        'Venture', 'Description', 'Payment Reference', 'Date'
    ])

    for t in transactions:
        writer.writerow([
            t.reference,
            t.get_transaction_type_display(),
            float(t.amount),
            float(t.balance_after),
            t.venture.title if t.venture else '',
            t.description or '',
            t.payment_reference or '',
            t.created_at.strftime('%Y-%m-%d %H:%M') if t.created_at else '',
        ])

    return response


@login_required
@partner_required
def export_my_withdrawals_csv(request):
    """Export user's withdrawals to CSV"""
    withdrawals = TokenWithdrawal.objects.filter(partner=request.user).order_by('-created_at')

    response = HttpResponse(content_type='text/csv')
    filename = f"my_withdrawals_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)
    writer.writerow([
        'Reference', 'Tokens (SRT)', 'Fee (NGN)', 'Amount (NGN)',
        'Bank Name', 'Account Number', 'Account Name',
        'Status', 'Created Date', 'Processed Date', 'Completed Date'
    ])

    for w in withdrawals:
        writer.writerow([
            w.reference,
            float(w.tokens),
            float(w.fee) if w.fee else 0,
            float(w.amount_ngn),
            w.get_bank_name_display(),
            w.account_number,
            w.account_name,
            w.get_status_display(),
            w.created_at.strftime('%Y-%m-%d %H:%M') if w.created_at else '',
            w.processed_at.strftime('%Y-%m-%d %H:%M') if w.processed_at else '',
            w.completed_at.strftime('%Y-%m-%d %H:%M') if w.completed_at else '',
        ])

    return response
