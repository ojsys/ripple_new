from django.contrib import admin
from django.utils.html import format_html
from django.urls import reverse
from django.http import HttpResponse
from django.utils import timezone
import csv
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from .models import (
    TokenPackage, PartnerCapitalAccount, Venture,
    VentureInvestment, SRTTransaction, TokenPurchase, TokenWithdrawal
)
from .email_utils import (
    send_withdrawal_approved_to_user,
    send_withdrawal_completed_to_user,
    send_withdrawal_rejected_to_user,
)


@admin.register(TokenPackage)
class TokenPackageAdmin(admin.ModelAdmin):
    list_display = ['name', 'tokens', 'bonus_tokens', 'total_tokens_display',
                    'price_ngn', 'price_usd', 'is_active', 'is_featured', 'order']
    list_filter = ['is_active', 'is_featured']
    list_editable = ['is_active', 'is_featured', 'order']
    search_fields = ['name']
    ordering = ['order', 'price_ngn']

    fieldsets = (
        (None, {
            'fields': ('name', 'description')
        }),
        ('Token Details', {
            'fields': ('tokens', 'bonus_tokens', 'price_ngn', 'price_usd')
        }),
        ('Display', {
            'fields': ('is_active', 'is_featured', 'order')
        }),
    )

    def total_tokens_display(self, obj):
        return f"{obj.total_tokens} SRT"
    total_tokens_display.short_description = "Total Tokens"


@admin.register(PartnerCapitalAccount)
class PartnerCapitalAccountAdmin(admin.ModelAdmin):
    list_display = ['partner_name', 'partner_id', 'token_balance', 'available_tokens_display',
                    'locked_tokens', 'total_investment_value_ngn', 'updated_at']
    list_filter = ['partner__partner_profile__accreditation_status']
    search_fields = ['partner__email', 'partner__first_name', 'partner__last_name',
                     'partner__partner_profile__partner_id']
    readonly_fields = ['partner', 'total_tokens_purchased', 'total_tokens_invested',
                       'total_tokens_earned', 'created_at', 'updated_at']

    fieldsets = (
        ('Partner', {
            'fields': ('partner',)
        }),
        ('Token Balance', {
            'fields': ('token_balance', 'locked_tokens')
        }),
        ('Statistics', {
            'fields': ('total_tokens_purchased', 'total_tokens_invested',
                       'total_tokens_earned', 'total_investment_value_ngn')
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )

    def partner_name(self, obj):
        return obj.partner.get_full_name() or obj.partner.email
    partner_name.short_description = "Partner"

    def partner_id(self, obj):
        try:
            return obj.partner.partner_profile.partner_id
        except:
            return "-"
    partner_id.short_description = "Partner ID"

    def available_tokens_display(self, obj):
        return f"{obj.available_tokens:.2f} SRT"
    available_tokens_display.short_description = "Available"


@admin.register(Venture)
class VentureAdmin(admin.ModelAdmin):
    list_display = ['title', 'industry', 'stage', 'risk_level', 'status',
                    'funding_progress', 'investor_count', 'is_featured']
    list_filter = ['status', 'stage', 'risk_level', 'industry', 'is_featured']
    list_editable = ['status', 'is_featured']
    search_fields = ['title', 'description', 'industry']
    prepopulated_fields = {'slug': ('title',)}
    readonly_fields = ['amount_raised', 'created_at', 'updated_at']

    fieldsets = (
        (None, {
            'fields': ('title', 'slug', 'short_description', 'description', 'image')
        }),
        ('Funding', {
            'fields': ('funding_goal', 'minimum_investment', 'maximum_investment', 'amount_raised')
        }),
        ('Investment Terms', {
            'fields': ('expected_return_rate', 'investment_duration_months')
        }),
        ('Classification', {
            'fields': ('stage', 'risk_level', 'industry', 'status')
        }),
        ('Dates', {
            'fields': ('start_date', 'end_date')
        }),
        ('Display', {
            'fields': ('is_featured', 'order', 'founder')
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )

    def funding_progress(self, obj):
        percent = obj.percent_funded
        color = 'success' if percent >= 100 else 'warning' if percent >= 50 else 'secondary'
        return format_html(
            '<div style="width:100px; background:#e9ecef; border-radius:4px;">'
            '<div style="width:{}%; background:var(--bs-{}); height:20px; border-radius:4px; text-align:center; color:white; font-size:12px; line-height:20px;">'
            '{:.1f}%</div></div>',
            min(percent, 100), color, percent
        )
    funding_progress.short_description = "Progress"


@admin.register(VentureInvestment)
class VentureInvestmentAdmin(admin.ModelAdmin):
    list_display = ['reference', 'partner_name', 'venture', 'tokens_invested',
                    'expected_return', 'status', 'maturity_date', 'created_at']
    list_filter = ['status', 'venture', 'created_at']
    search_fields = ['reference', 'partner__email', 'partner__first_name',
                     'partner__last_name', 'venture__title']
    readonly_fields = ['reference', 'partner', 'venture', 'account', 'tokens_invested',
                       'expected_return', 'investment_date', 'created_at', 'updated_at']
    date_hierarchy = 'created_at'

    fieldsets = (
        ('Investment Details', {
            'fields': ('reference', 'partner', 'venture', 'account')
        }),
        ('Tokens', {
            'fields': ('tokens_invested', 'expected_return', 'actual_return')
        }),
        ('Status', {
            'fields': ('status', 'maturity_date')
        }),
        ('Notes', {
            'fields': ('notes',)
        }),
        ('Timestamps', {
            'fields': ('investment_date', 'created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )

    def partner_name(self, obj):
        return obj.partner.get_full_name() or obj.partner.email
    partner_name.short_description = "Partner"


@admin.register(SRTTransaction)
class SRTTransactionAdmin(admin.ModelAdmin):
    list_display = ['reference', 'partner_name', 'transaction_type', 'amount_display',
                    'balance_after', 'venture', 'created_at']
    list_filter = ['transaction_type', 'created_at']
    search_fields = ['reference', 'account__partner__email',
                     'account__partner__first_name', 'account__partner__last_name']
    readonly_fields = ['account', 'transaction_type', 'amount', 'balance_after',
                       'venture', 'token_package', 'reference', 'payment_reference',
                       'payment_method', 'created_at']
    date_hierarchy = 'created_at'

    fieldsets = (
        ('Transaction', {
            'fields': ('reference', 'transaction_type', 'description')
        }),
        ('Account', {
            'fields': ('account', 'amount', 'balance_after')
        }),
        ('Related', {
            'fields': ('venture', 'token_package')
        }),
        ('Payment', {
            'fields': ('payment_reference', 'payment_method'),
            'classes': ('collapse',)
        }),
        ('Timestamp', {
            'fields': ('created_at',)
        }),
    )

    def partner_name(self, obj):
        return obj.account.partner.get_full_name() or obj.account.partner.email
    partner_name.short_description = "Partner"

    def amount_display(self, obj):
        try:
            amount = float(obj.amount)
        except (ValueError, TypeError):
            return obj.amount  # Failsafe for non-numeric data

        if amount >= 0:
            formatted_amount = f"+{amount:.2f}"
            return format_html('<span style="color:green;">{}</span>', formatted_amount)
        else:
            formatted_amount = f"{amount:.2f}"
            return format_html('<span style="color:red;">{}</span>', formatted_amount)
    amount_display.short_description = "Amount"

    actions = ['export_to_excel', 'export_to_csv']

    def export_to_excel(self, request, queryset):
        """Export selected transactions to Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Transactions"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        headers = [
            'Reference', 'Partner Name', 'Partner Email', 'Type', 'Amount (SRT)',
            'Balance After', 'Venture', 'Description', 'Payment Reference', 'Date'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        for row_num, t in enumerate(queryset.order_by('-created_at'), 2):
            row_data = [
                t.reference,
                t.account.partner.get_full_name() or t.account.partner.email,
                t.account.partner.email,
                t.get_transaction_type_display(),
                float(t.amount),
                float(t.balance_after),
                t.venture.title if t.venture else '',
                t.description or '',
                t.payment_reference or '',
                t.created_at.strftime('%Y-%m-%d %H:%M') if t.created_at else '',
            ]
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border

        # Set column widths
        widths = [18, 20, 25, 18, 15, 15, 25, 40, 20, 18]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"transactions_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response.write(buffer.getvalue())

        return response
    export_to_excel.short_description = "Export selected to Excel"

    def export_to_csv(self, request, queryset):
        """Export selected transactions to CSV"""
        response = HttpResponse(content_type='text/csv')
        filename = f"transactions_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        writer = csv.writer(response)
        writer.writerow([
            'Reference', 'Partner Name', 'Partner Email', 'Type', 'Amount (SRT)',
            'Balance After', 'Venture', 'Description', 'Payment Reference', 'Date'
        ])

        for t in queryset.order_by('-created_at'):
            writer.writerow([
                t.reference,
                t.account.partner.get_full_name() or t.account.partner.email,
                t.account.partner.email,
                t.get_transaction_type_display(),
                float(t.amount),
                float(t.balance_after),
                t.venture.title if t.venture else '',
                t.description or '',
                t.payment_reference or '',
                t.created_at.strftime('%Y-%m-%d %H:%M') if t.created_at else '',
            ])

        return response
    export_to_csv.short_description = "Export selected to CSV"


@admin.register(TokenPurchase)
class TokenPurchaseAdmin(admin.ModelAdmin):
    list_display = ['paystack_reference', 'partner_name', 'package', 'tokens',
                    'bonus_tokens', 'amount_ngn', 'status', 'created_at']
    list_filter = ['status', 'package', 'created_at']
    search_fields = ['paystack_reference', 'partner__email',
                     'partner__first_name', 'partner__last_name']
    readonly_fields = ['partner', 'account', 'package', 'tokens', 'bonus_tokens',
                       'amount_ngn', 'amount_usd', 'paystack_reference',
                       'created_at', 'completed_at']
    date_hierarchy = 'created_at'

    fieldsets = (
        ('Purchase', {
            'fields': ('paystack_reference', 'partner', 'account', 'package')
        }),
        ('Tokens', {
            'fields': ('tokens', 'bonus_tokens')
        }),
        ('Payment', {
            'fields': ('amount_ngn', 'amount_usd', 'status')
        }),
        ('Timestamps', {
            'fields': ('created_at', 'completed_at')
        }),
    )

    def partner_name(self, obj):
        return obj.partner.get_full_name() or obj.partner.email
    partner_name.short_description = "Partner"

    actions = ['mark_successful']

    def mark_successful(self, request, queryset):
        for purchase in queryset.filter(status='pending'):
            purchase.complete_purchase()
        self.message_user(request, f"Marked {queryset.count()} purchases as successful")
    mark_successful.short_description = "Mark selected purchases as successful"


@admin.register(TokenWithdrawal)
class TokenWithdrawalAdmin(admin.ModelAdmin):
    list_display = ['reference', 'partner_name', 'tokens', 'amount_ngn',
                    'bank_name', 'status', 'created_at']
    list_filter = ['status', 'bank_name', 'created_at']
    search_fields = ['reference', 'partner__email', 'partner__first_name',
                     'partner__last_name', 'account_number', 'account_name']
    readonly_fields = ['partner', 'account', 'tokens', 'amount_ngn',
                       'bank_name', 'account_number', 'account_name',
                       'reference', 'created_at', 'processed_at', 'completed_at']
    date_hierarchy = 'created_at'

    fieldsets = (
        ('Withdrawal Request', {
            'fields': ('reference', 'partner', 'account', 'status')
        }),
        ('Token Details', {
            'fields': ('tokens', 'amount_ngn')
        }),
        ('Bank Details', {
            'fields': ('bank_name', 'account_number', 'account_name')
        }),
        ('Processing', {
            'fields': ('admin_notes', 'processed_by', 'payment_reference')
        }),
        ('Timestamps', {
            'fields': ('created_at', 'processed_at', 'completed_at'),
            'classes': ('collapse',)
        }),
    )

    def partner_name(self, obj):
        return obj.partner.get_full_name() or obj.partner.email
    partner_name.short_description = "Partner"

    actions = ['approve_withdrawals', 'reject_withdrawals', 'mark_completed', 'export_to_excel', 'export_to_csv']

    def approve_withdrawals(self, request, queryset):
        count = 0
        for withdrawal in queryset.filter(status='pending'):
            if withdrawal.approve(request.user):
                count += 1
                # Send email notification to user
                send_withdrawal_approved_to_user(withdrawal)
        self.message_user(request, f"Approved {count} withdrawal requests and sent email notifications")
    approve_withdrawals.short_description = "Approve selected withdrawals"

    def reject_withdrawals(self, request, queryset):
        count = 0
        for withdrawal in queryset.filter(status='pending'):
            if withdrawal.reject(request.user, "Rejected by admin"):
                count += 1
                # Send email notification to user
                send_withdrawal_rejected_to_user(withdrawal)
        self.message_user(request, f"Rejected {count} withdrawal requests and sent email notifications")
    reject_withdrawals.short_description = "Reject selected withdrawals"

    def mark_completed(self, request, queryset):
        count = 0
        for withdrawal in queryset.filter(status__in=['approved', 'processing']):
            if withdrawal.complete():
                count += 1
                # Send email notification to user
                send_withdrawal_completed_to_user(withdrawal)
        self.message_user(request, f"Marked {count} withdrawals as completed and sent email notifications")
    mark_completed.short_description = "Mark selected as completed"

    def export_to_excel(self, request, queryset):
        """Export selected withdrawals to Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Withdrawals"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        headers = [
            'Reference', 'Partner Name', 'Partner Email', 'Tokens (SRT)', 'Fee (NGN)',
            'Amount (NGN)', 'Bank Name', 'Account Number', 'Account Name',
            'Status', 'Created Date', 'Processed Date', 'Completed Date', 'Admin Notes'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        for row_num, w in enumerate(queryset.order_by('-created_at'), 2):
            row_data = [
                w.reference,
                w.partner.get_full_name() or w.partner.email,
                w.partner.email,
                float(w.tokens),
                float(w.fee) if w.fee else 0,
                float(w.amount_ngn),
                w.get_bank_name_display(),
                w.account_number,
                w.account_name,
                w.get_status_display(),
                w.created_at.strftime('%Y-%m-%d %H:%M') if w.created_at else '',
                w.processed_at.strftime('%Y-%m-%d %H:%M') if w.processed_at else '',
                w.completed_at.strftime('%Y-%m-%d %H:%M') if w.completed_at else '',
                w.admin_notes or '',
            ]
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border

        # Set column widths
        widths = [15, 20, 25, 12, 12, 15, 25, 15, 25, 12, 18, 18, 18, 30]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"withdrawals_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response.write(buffer.getvalue())

        return response
    export_to_excel.short_description = "Export selected to Excel"

    def export_to_csv(self, request, queryset):
        """Export selected withdrawals to CSV"""
        response = HttpResponse(content_type='text/csv')
        filename = f"withdrawals_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        writer = csv.writer(response)
        writer.writerow([
            'Reference', 'Partner Name', 'Partner Email', 'Tokens (SRT)', 'Fee (NGN)',
            'Amount (NGN)', 'Bank Name', 'Account Number', 'Account Name',
            'Status', 'Created Date', 'Processed Date', 'Completed Date', 'Admin Notes'
        ])

        for w in queryset.order_by('-created_at'):
            writer.writerow([
                w.reference,
                w.partner.get_full_name() or w.partner.email,
                w.partner.email,
                float(w.tokens),
                float(w.fee) if w.fee else 0,
                float(w.amount_ngn),
                w.get_bank_name_display(),
                w.account_number,
                w.account_name,
                w.get_status_display(),
                w.created_at.strftime('%Y-%m-%d %H:%M') if w.created_at else '',
                w.processed_at.strftime('%Y-%m-%d %H:%M') if w.processed_at else '',
                w.completed_at.strftime('%Y-%m-%d %H:%M') if w.completed_at else '',
                w.admin_notes or '',
            ])

        return response
    export_to_csv.short_description = "Export selected to CSV"
